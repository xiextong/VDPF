from tqdm import tqdm
import os
from openpyxl import Workbook, load_workbook
import numpy as np
import pandas as pd
import SimpleITK as sitk
from scipy.ndimage import zoom

def get_box(mask):
    """
    :param mask:  array
    :return:
    """
    indexx = np.where(mask > 0.)
    dim0min, dim0max, dim1min, dim1max, dim2min, dim2max = [np.min(indexx[0]), np.max(indexx[0]),
                                                            np.min(indexx[1]), np.max(indexx[1]),
                                                            np.min(indexx[2]), np.max(indexx[2])]
    bbox = [dim0min, dim0max, dim1min, dim1max, dim2min, dim2max]
    return bbox

def readIMG_from_sitk(filename):
    """
    read mhd/NIFTI image
    :param filename:
    :return:
    scan, spacing, origin, transfmat, axesOrder
    """
    itkimage = sitk.ReadImage(filename)
    scan = sitk.GetArrayFromImage(itkimage)
    spacing = itkimage.GetSpacing()
    origin = itkimage.GetOrigin()  # world coordinates of origin
    transfmat = itkimage.GetDirection()  # 3D rotation matrix
    axesOrder = ['axial', 'coronal', 'sagittal']  #
    return scan, spacing, origin, transfmat, axesOrder


def writeIMG_as_sitk(filename, scan, spacing, origin, transfmat):
    """
    :param filename:
    :param scan: axis of scan must be [axial,coronal,sagittal], means [z，y，x]
    :param spacing:
    :param origin:
    :param transfmat:
    :return:
    """

    itkim = sitk.GetImageFromArray(scan, isVector=False)  # 3D image
    itkim.SetSpacing(spacing)  # voxelsize
    itkim.SetOrigin(origin)  # world coordinates of origin
    itkim.SetDirection(transfmat)  # 3D rotation matrix
    sitk.WriteImage(itkim, filename, False)

def writesimple(path, list, index, column, sheetname='Sheet1'):
    '''
    :param path: save path
    :param list: data
    :param index: row
    :param column: column
    :param sheetname: sheetname
    :return:
    example: writesimple(filename, '0train_1test', 1, 2)
    '''
    if os.path.exists(path):
        bg = load_workbook(path)
        sheets = bg.sheetnames
        if sheetname in sheets:
            sheet = bg[str(sheetname)]
            sheet.cell(index, column, list)
            bg.save(path)
            bg.close()
        else:
            sheet = bg.create_sheet(str(sheetname))
            sheet = bg[str(sheetname)]
            sheet.cell(index, column, list)
            bg.save(path)
            bg.close()
    else:
        bg1 = Workbook()
        bg1.active
        bg1.save(path)
        bg1.close()
        bg = load_workbook(path)
        sheet = bg.create_sheet(str(sheetname))
        # sheet = bg[str(sheetname)]
        sheet.cell(index, column, list)
        bg.save(path)
        bg.close()

def judgedir(path):
    '''
    Args:
        path: str，
    '''
    sep = os.sep
    p_path = path.split(sep)
    fullpath = p_path[0]
    for n, ff in enumerate(p_path[1:]):
        fullpath = fullpath + sep + ff
        if os.path.exists(fullpath):
            pass
        else:
            os.makedirs(path)

def segmenting_dvt(mri, roi,bbox=None):

    dim00, dim11, dim22 = np.where(roi == 2)
    dim111 = sorted(set(dim11))
    if len(dim111) != 2:
        return [0]
    else:
        roi[:, dim111[0], :] = 2
        roi[:, dim111[1], :] = 2

        A_MRI = mri[:, :dim111[0], :]
        B_MRI = mri[:, dim111[0]:dim111[1], :]
        A_ROI = roi[:, :dim111[0], :]
        B_ROI = roi[:, dim111[0]:dim111[1], :]
        if bbox is not None:
            A_bbox = bbox[:, :dim111[0], :]
            B_bbox = bbox[:, dim111[0]:dim111[1], :]
            return A_MRI, B_MRI, A_ROI, B_ROI, A_bbox, B_bbox
        else:
            return A_MRI, B_MRI, A_ROI, B_ROI


if __name__ == '_1_seg_big_data__':
    ''' partitioning the data '''
    rootpath = r'..\Data\ori_data'
    seg_result_root = r'..\Data\seg_big_data'
    judgedir(seg_result_root)
    excelpath = os.path.join(r'..\Data\label.xlsx')
    #read excel
    excel_data = pd.read_excel(excelpath)
    pth_num = 0
    data_num = 0

    for file in os.listdir(rootpath):
        A1 = 0
        A2 = 0
        B1 = 0
        B2 = 0
        if '_' not in file and 'nii' in file:
            pth_name = file.split('.')[0]
            mri_path = os.path.join(rootpath, file)
            roi_path = mri_path.replace('.nii.gz', '_ROI2.nii.gz')
            bbox_path = mri_path.replace('.nii.gz', '_Rectf.nii.gz')

            mri, spacing_mri, origin_mri, transfmat_mri, axesOrder_mri = readIMG_from_sitk(mri_path)
            roi, spacing_roi, origin_roi, transfmat_roi, axesOrder_roi = readIMG_from_sitk(roi_path)
            bbox, spacing_bbox, origin_bbox, transfmat_bbox, axesOrder_bbox = readIMG_from_sitk(bbox_path)

            seg_result = segmenting_dvt(mri, roi)
            if  len(seg_result) != 6:
                print(pth_name + 'is wrong......')
                continue
            else:
                A_MRI, B_MRI, A_ROI, B_ROI, A_bbox, B_bbox = seg_result

            pth_name1 = int(pth_name)
            middle_all = int(excel_data[excel_data.num == pth_name1].middle)

            ############## part of B ######################
            B_MRI_R = None #right
            B_MRI_L = None #left
            B_bbox_L = None
            B_bbox_R = None
            if np.sum(B_ROI) != 0:
                [_, _, _, _, dim2min_b, dim2max_b] = get_box(B_ROI)
                if np.sum(B_ROI[:, :, middle_all]) != 0:
                    print(pth_name, '中线上有值，不正常，请检查！！！！！！！！！！！！！！！！！！！')
                    continue
                if np.sum(B_ROI[:, :, :middle_all]) != 0 and np.sum(B_ROI[:, :, middle_all + 1:]) != 0:  # both sides
                    B1 = 1
                    B2 = 1
                    B_MRI_R = B_MRI[:, :, :middle_all]
                    B_ROI_R = B_ROI[:, :, :middle_all]

                    B_MRI_L = B_MRI[:, :, middle_all:]
                    B_ROI_L = B_ROI[:, :, middle_all:]

                    B_bbox_R = B_bbox[:, :, :middle_all]
                    B_bbox_L = B_bbox[:, :, middle_all:]


                elif middle_all <= dim2min_b: #one side
                    B2 = 1
                    B_MRI_L = B_MRI[:, :, middle_all:]
                    B_ROI_L = B_ROI[:, :, middle_all:]
                    B_bbox_L = B_bbox[:, :, middle_all:]

                elif middle_all > dim2max_b:
                    B1 = 1
                    B_ROI_R = B_ROI[:, :, :middle_all]
                    B_MRI_R = B_MRI[:, :, :middle_all]
                    B_bbox_R = B_bbox[:, :, :middle_all]
                else:
                    continue

                if B_MRI_R is not None:
                    resB_MRI_R = os.path.join(seg_result_root, pth_name + '_B_R_MRI' + '.nii.gz')
                    resB_ROI_R = resB_MRI_R.replace('MRI', 'ROI')
                    writeIMG_as_sitk(resB_MRI_R, B_MRI_R, spacing_mri, origin_mri, transfmat_mri)
                    writeIMG_as_sitk(resB_ROI_R, B_ROI_R, spacing_roi, origin_roi, transfmat_roi)
                    resB_bbox_R = resB_MRI_R.replace('.nii.gz', '_Rectf.nii.gz')
                    writeIMG_as_sitk(resB_bbox_R, B_bbox_R, spacing_bbox, origin_bbox, transfmat_bbox)
                if B_MRI_L is not None:
                    resB_MRI_L = os.path.join(seg_result_root, pth_name + '_B_L_MRI' + '.nii.gz')
                    resB_ROI_L = resB_MRI_L.replace('MRI', 'ROI')
                    writeIMG_as_sitk(resB_MRI_L, B_MRI_L, spacing_mri, origin_mri, transfmat_mri)
                    writeIMG_as_sitk(resB_ROI_L, B_ROI_L, spacing_roi, origin_roi, transfmat_roi)
                    resB_bbox_L = resB_MRI_L.replace('.nii.gz', '_Rectf.nii.gz')
                    writeIMG_as_sitk(resB_bbox_L, B_bbox_L, spacing_bbox, origin_bbox, transfmat_bbox)

            ############## part of A ######################
            A_MRI_L = None
            A_MRI_R = None
            A_bbox_L = None
            A_bbox_R = None
            if A_MRI.shape[1]!=0:
                if np.sum(A_ROI) != 0:
                    [_, _, _, _, dim2min_a, dim2max_a] = get_box(A_ROI)

                    if np.sum(A_ROI[:, :, :middle_all]) != 0 and np.sum(A_ROI[:, :, middle_all + 1:]) != 0:   # both sides
                        A1 = 1
                        A2 = 1
                        A_MRI_R = A_MRI[:, :, :middle_all]
                        A_ROI_R = A_ROI[:, :, :middle_all]
                        A_MRI_L = A_MRI[:, :, middle_all:]
                        A_ROI_L = A_ROI[:, :, middle_all:]

                        A_bbox_R = A_bbox[:, :, :middle_all]
                        A_bbox_L = A_bbox[:, :, middle_all:]

                    elif middle_all <= dim2min_a: #one side
                        A2 = 1
                        A_MRI_L = A_MRI[:, :, middle_all:]
                        A_ROI_L = A_ROI[:, :, middle_all:]
                        A_bbox_L = A_bbox[:, :, middle_all:]

                    elif middle_all >= dim2max_a:
                        A1 = 1
                        A_ROI_R = A_ROI[:, :, :middle_all]
                        A_MRI_R = A_MRI[:, :, :middle_all]
                        A_bbox_R = A_bbox[:, :, :middle_all]
                    else:
                        continue

                    if A_MRI_R is not None:
                        resA_MRI_R = os.path.join(seg_result_root, pth_name + '_A_R_MRI' + '.nii.gz')
                        resA_ROI_R = resA_MRI_R.replace('MRI', 'ROI')
                        writeIMG_as_sitk(resA_MRI_R, A_MRI_R, spacing_mri, origin_mri, transfmat_mri)
                        writeIMG_as_sitk(resA_ROI_R, A_ROI_R, spacing_roi, origin_roi, transfmat_roi)
                        resA_bbox_R = resA_MRI_R.replace('.nii.gz', '_Rectf.nii.gz')
                        writeIMG_as_sitk(resA_bbox_R, A_bbox_R, spacing_bbox, origin_bbox, transfmat_bbox)
                    if A_MRI_L is not None:
                        resA_MRI_L = os.path.join(seg_result_root, pth_name + '_A_L_MRI' + '.nii.gz')
                        resA_ROI_L = resA_MRI_L.replace('MRI', 'ROI')
                        writeIMG_as_sitk(resA_MRI_L, A_MRI_L, spacing_mri, origin_mri, transfmat_mri)
                        writeIMG_as_sitk(resA_ROI_L, A_ROI_L, spacing_roi, origin_roi, transfmat_roi)
                        resA_bbox_L = resA_MRI_L.replace('.nii.gz', '_Rectf.nii.gz')
                        writeIMG_as_sitk(resA_bbox_L, A_bbox_L, spacing_bbox, origin_bbox, transfmat_bbox)
            data_num = data_num + A1 + A2 + B1 + B2
            pth_num = pth_num + 1
    print('data_num  ', data_num)
    print('pth_num  ', pth_num)

if __name__ == '_2_seg_big_data_spacing':
    ''' spacing -> 1  '''
    rootpath = r'..\Data\seg_big_data'
    seg_result_root = r'..\Data\seg_big_data_spacing1_1'
    judgedir(seg_result_root)

    for file in os.listdir(rootpath):
        full_name = file.rsplit('_', 1)[0]
        if '_MRI' in file:
            mri_path = os.path.join(rootpath, file)
            roi_path = mri_path.replace('MRI', 'ROI')

            mri, spacing_mri, origin_mri, transfmat_mri, axesOrder_mri = readIMG_from_sitk(mri_path)
            roi, spacing_roi, origin_roi, transfmat_roi, axesOrder_roi = readIMG_from_sitk(roi_path)
            print(full_name)

            mri = zoom(mri, (1 / spacing_mri[0], 1, 1 / spacing_mri[2]), order=3)
            roi = zoom(roi, (1 / spacing_mri[0], 1, 1 / spacing_mri[2]), order=3)
            new_spacing = (1.0, spacing_mri[1], 1.0)

            writeIMG_as_sitk(os.path.join(seg_result_root, file), mri, new_spacing, origin_mri, transfmat_mri)
            writeIMG_as_sitk(os.path.join(seg_result_root, file.replace('MRI', 'ROI')), roi, new_spacing, origin_roi,transfmat_roi)


if __name__ == '_3_seg_big_data_spacing_224_':
    '''resize'''
    rootpath = r'..\Data\seg_big_data_spacing1_1'
    seg_result_root = r'..\Data\seg_big_data_spacing1_1_224'
    judgedir(seg_result_root)
    target_size = [224, 224]

    tqdm_bar = tqdm(os.listdir(rootpath), desc='3D resize')
    for file in tqdm_bar:
        full_name = file.rsplit('_', 1)[0]
        if '_MRI' in file:
            # pth_name = file.split('_')[0]
            mri_path = os.path.join(rootpath, file)
            roi_path = mri_path.replace('MRI', 'ROI')
            mri, spacing_mri, origin_mri, transfmat_mri, axesOrder_mri = readIMG_from_sitk(mri_path)
            roi, spacing_roi, origin_roi, transfmat_roi, axesOrder_roi = readIMG_from_sitk(roi_path)
            zz,yy,xx = mri.shape
            mri = zoom(mri, (224/zz, 1, 224/xx), order=3)
            roi = zoom(roi,(224/zz, 1, 224/xx), order=3)
            new_spacing = ( spacing_mri[0]*224/zz, spacing_mri[1], spacing_mri[2]*224/xx)
            writeIMG_as_sitk(os.path.join(seg_result_root, file), mri, new_spacing, origin_mri, transfmat_mri)
            writeIMG_as_sitk(os.path.join(seg_result_root, file.replace('MRI', 'ROI')), roi, new_spacing, origin_roi,transfmat_roi)

if __name__ == '_4_seg_big_data_spacing_224_cut_slice_nor_':
    '''
    adjust window width and level
    3d to 2d
    Normalization
    '''
    rootpath = r'..\Data\seg_big_data_spacing1_1_224'

    for file in os.listdir(rootpath):
        full_name = file.rsplit('_', 1)[0]
        if '_MRI' in file:
            print(full_name)
            mri_path = os.path.join(rootpath, file)
            roi_path = mri_path.replace('MRI', 'ROI')

            mri, spacing_mri, origin_mri, transfmat_mri, axesOrder_mri = readIMG_from_sitk(mri_path)
            roi, spacing_roi, origin_roi, transfmat_roi, axesOrder_roi = readIMG_from_sitk(roi_path)

            # adjust window width and level
            cut_thre1 = np.percentile(mri, 99.91)
            cut_thre2 = np.percentile(mri, 0.09)
            mri[mri >= cut_thre1] = cut_thre1
            mri[mri <= cut_thre2] = cut_thre2

            # 3d to 2d
            for yyy in range(mri.shape[1]):
                img_slice = mri[:, yyy, :]
                mask_slice = roi[:, yyy, :]
                # Normalization
                img_slice = (img_slice - np.min(img_slice)) / (np.max(img_slice) - np.min(img_slice))
                final_data = {'img': img_slice, 'mask': mask_slice}











